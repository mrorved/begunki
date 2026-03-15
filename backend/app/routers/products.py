import os
import io
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, or_, func
from typing import Optional
from app.database import get_db
from app.models import Product
from app.auth import get_current_user, require_admin
import openpyxl

router = APIRouter()

def get_photos_dir() -> str:
    """Photos dir: env var PHOTOS_DIR, or <project_root>/photos"""
    # PHOTOS_DIR env set by main.py at startup
    d = os.getenv("PHOTOS_DIR")
    if d:
        os.makedirs(d, exist_ok=True)
        return d
    # Fallback: go up from routers/ -> app/ -> backend/ -> project root
    base = os.path.dirname(  # project root
               os.path.dirname(  # backend/
                   os.path.dirname(  # app/
                       os.path.dirname(  # routers/
                           os.path.abspath(__file__)))))
    photos = os.path.join(base, "photos")
    os.makedirs(photos, exist_ok=True)
    return photos


@router.get("")
async def get_products(
    search: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    manufacturer: Optional[str] = Query(None),
    in_stock: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    q = select(Product)
    if search:
        like = f"%{search}%"
        q = q.where(or_(
            Product.name.ilike(like),
            Product.code.ilike(like),
            Product.manufacturer.ilike(like),
            Product.type.ilike(like),
        ))
    if type:
        q = q.where(Product.type == type)
    if manufacturer:
        q = q.where(Product.manufacturer == manufacturer)
    if in_stock == '1':
        q = q.where(Product.stock > 0)

    total_q = select(func.count()).select_from(q.subquery())
    total = (await db.execute(total_q)).scalar()

    q = q.order_by(Product.name).offset(skip).limit(limit)
    rows = (await db.execute(q)).scalars().all()

    return {
        "total": total,
        "items": [_to_dict(p) for p in rows],
    }


@router.get("/filters")
async def get_filters(db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    types = (await db.execute(select(Product.type).distinct().order_by(Product.type))).all()
    manufacturers = (await db.execute(select(Product.manufacturer).distinct().order_by(Product.manufacturer))).all()
    return {
        "types": [r[0] for r in types if r[0]],
        "manufacturers": [r[0] for r in manufacturers if r[0]],
    }


@router.get("/by-code/{code}")
async def get_by_code(code: str, db: AsyncSession = Depends(get_db), _=Depends(get_current_user)):
    result = await db.execute(select(Product).where(Product.code == code))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Товар не найден")
    return _to_dict(product)


@router.post("/import")
async def import_price(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_admin),
):
    content = await file.read()
    if not content:
        raise HTTPException(400, "Файл пустой")

    # Try openpyxl first (xlsx), then xlrd for old xls
    wb = None
    last_error = ""

    # Attempt 1: openpyxl with read_only=False
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    except Exception as e1:
        last_error = str(e1)

    # Attempt 2: openpyxl keep_vba
    if wb is None:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, keep_vba=True)
        except Exception as e2:
            last_error = str(e2)

    if wb is None:
        raise HTTPException(400, f"Не удалось открыть файл. Убедитесь что файл сохранён в формате .xlsx. Ошибка: {last_error}")

    ws = wb.active
    photos_dir = get_photos_dir()

    # Find header row — look for row containing "Наименование"
    data_start_row = 2  # default: second row
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row and any("аименован" in str(c or "") for c in row):
            data_start_row = i + 1
            break

    # Clear all products
    await db.execute(delete(Product))
    await db.flush()

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not row or not row[0]:
            continue

        # Skip non-numeric first column (e.g., repeated headers)
        try:
            int(str(row[0]).strip())
        except (ValueError, TypeError):
            continue

        try:
            name = str(row[1] or "").strip()
            code = str(row[2] or "").strip()
            grd_code = str(row[3] or "").strip()
            price = float(row[4] or 0)
            package = int(float(row[5] or 1))
            stock = int(float(row[6] or 0))
            ptype = str(row[7] or "").strip()
            manufacturer = str(row[8] or "").strip()

            if not code or not name:
                skipped += 1
                continue

            # Detect photo
            photo = None
            for ext in (".jpg", ".jpeg", ".png", ".webp", ".JPG", ".JPEG", ".PNG"):
                if os.path.exists(os.path.join(photos_dir, f"{code}{ext}")):
                    photo = f"{code}{ext}"
                    break

            db.add(Product(
                name=name, code=code, grd_code=grd_code,
                price=price, package=package, stock=stock,
                type=ptype, manufacturer=manufacturer, photo=photo,
            ))
            imported += 1

        except Exception:
            skipped += 1
            continue

    await db.commit()
    return {"imported": imported, "skipped": skipped}


@router.post("/upload-photo")
async def upload_photo(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_admin),
):
    """Upload product photo. Filename must match product code (e.g. 00015574.jpg)."""
    photos_dir = get_photos_dir()
    dest = os.path.join(photos_dir, file.filename)
    data = await file.read()
    with open(dest, "wb") as f:
        f.write(data)

    # Extract product code from filename (strip extension)
    code = os.path.splitext(file.filename)[0]

    # Update photo field in DB:
    # Try exact match first, then zero-padded variants (e.g. 64 -> 00000064)
    product = None
    matched_code = None

    # Exact match
    result = await db.execute(select(Product).where(Product.code == code))
    product = result.scalar_one_or_none()
    if product:
        matched_code = code

    # Try zero-padded to 8 digits
    if not product:
        padded = code.zfill(8)
        result = await db.execute(select(Product).where(Product.code == padded))
        product = result.scalar_one_or_none()
        if product:
            matched_code = padded

    # Try other common padding lengths (5, 6, 7)
    if not product:
        for pad_len in (5, 6, 7):
            padded = code.zfill(pad_len)
            result = await db.execute(select(Product).where(Product.code == padded))
            product = result.scalar_one_or_none()
            if product:
                matched_code = padded
                break

    if product:
        product.photo = file.filename
        await db.commit()
        print(f"[PHOTO] Updated product {matched_code} -> {file.filename}")
    else:
        print(f"[PHOTO] WARNING: product with code '{code}' not found in DB (tried padded variants too)")

    return {
        "ok": True,
        "saved": file.filename,
        "code_extracted": code,
        "product_updated": product is not None
    }


def _to_dict(p: Product) -> dict:
    return {
        "id": p.id,
        "name": p.name,
        "code": p.code,
        "grd_code": p.grd_code,
        "price": p.price,
        "package": p.package,
        "stock": p.stock,
        "type": p.type,
        "manufacturer": p.manufacturer,
        "photo": p.photo,
    }
